from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference

file_path = 'table.xlsx'  # Замените на путь к вашему файлу
wb = load_workbook(file_path)
ws = wb.active

ws["B14"] = "Средняя зарплата по отделам:"
ws["C14"] = (ws[4][5].value + ws[9][5].value + ws[11][5].value) / 3
minimum = ws[2][5].value
maximum = 0
for i in range(2, 10):
    if i not in [4, 9]:
        minimum = min(minimum, ws[i][5].value)
        maximum = max(maximum, ws[i][5].value)
ws["B15"] = "Максимальная зарплата:"
ws["C15"] = maximum
ws["B16"] = "Минимальная зарплата:"
ws["C16"] = minimum

wb.save("table.xlsx")
