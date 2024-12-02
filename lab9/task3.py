from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference

file_path = 'table.xlsx'
wb = load_workbook(file_path)
ws = wb.active

departments = {}
for i in [4, 9, 11]:
    departments[ws[i][2].value.split('\n')[0]] = ws[i][5].value

summary_sheet = wb.create_sheet(title="Summary")
summary_sheet.append(["Отдел", "Сумма зарплаты"])
for dept, salary in departments.items():
    summary_sheet.append([dept, salary])
chart = PieChart()
chart.title = "Распределение зарплат по отделам"
labels = Reference(summary_sheet, min_col=1, min_row=2, max_row=len(departments) + 1)
data = Reference(summary_sheet, min_col=2, min_row=1, max_row=len(departments) + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
ws.add_chart(chart, "K1")

output_file_path = 'table.xlsx'  # Замените на нужный путь
wb.save(output_file_path)